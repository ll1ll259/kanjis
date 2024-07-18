from tkinter import *
from tkinter import messagebox
import pandas as pd
import numpy as np
from openpyxl import load_workbook
import os


def kanjis():
    n = int(num_kanjis.get())
    df = pd.read_excel("A:\Rubén\Librerías\Documentos\Carpeta Compartida Linux\漢字.xlsx")
    readings = ['次','くにょみ','おにょみ']
    aux = ['次','いみ','漢字',]
    new_kanji = df.loc[df['新しい']==1]
    df = df.loc[df['新しい']==0]
    df['random'] =  np.random.random(size=len(df))
    df.sort_values(by="random", inplace = True, ascending=False)


    if(new_kj.get()==1):

        if(excel_display.get()==0): 
            daily_kanji_readings = pd.concat([df[readings].head(n).reset_index(drop = True),new_kanji[readings]])
            daily_kanji = pd.concat([df[aux].head(n).reset_index(drop = True),new_kanji[aux]])
            daily_kanji_readings.to_excel('C:/Users/ruben/Desktop/毎日の漢字.xlsx',"lecturas",index = False)
            book = load_workbook('C:/Users/ruben/Desktop/毎日の漢字.xlsx')
            writer = pd.ExcelWriter('C:/Users/ruben/Desktop/毎日の漢字.xlsx', engine='openpyxl') 
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            daily_kanji.to_excel(writer, "漢字, いみ",index = False)
            writer.save()
            messagebox.showinfo('毎日の漢字プロガム','満たされた')              

        if(excel_display.get()==1): 
            daily_kanji_readings = pd.concat([df[readings].head(n).reset_index(drop = True),new_kanji[readings]])
            daily_kanji = pd.concat([df[aux].head(n).reset_index(drop = True),new_kanji[aux]])
            daily_kanji_readings.to_excel('C:/Users/ruben/Desktop/毎日の漢字.xlsx',"lecturas",index = False)
            book = load_workbook('C:/Users/ruben/Desktop/毎日の漢字.xlsx')
            writer = pd.ExcelWriter('C:/Users/ruben/Desktop/毎日の漢字.xlsx', engine='openpyxl') 
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            daily_kanji.to_excel(writer, "漢字, いみ",index = False)
            writer.save()
            os.system(r"start EXCEL.EXE C:\Users\ruben\Desktop\毎日の漢字.xlsx")
                    
    

    if(new_kj.get()==0):
        
        if(excel_display.get()==0): 
            daily_kanji_readings = df[readings].head(n).reset_index(drop = True)
            daily_kanji = df[aux].head(n).reset_index(drop = True)
            daily_kanji_readings.to_excel('C:/Users/ruben/Desktop/毎日の漢字.xlsx',"lecturas",index = False)
            book = load_workbook('C:/Users/ruben/Desktop/毎日の漢字.xlsx')
            writer = pd.ExcelWriter('C:/Users/ruben/Desktop/毎日の漢字.xlsx', engine='openpyxl') 
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            daily_kanji.to_excel(writer, "漢字, いみ",index = False)
            writer.save()
            messagebox.showinfo('毎日の漢字プロガム','満たされた')      
                    

        if(excel_display.get()==1): 
            daily_kanji_readings = df[readings].head(n).reset_index(drop = True)
            daily_kanji = df[aux].head(n).reset_index(drop = True)
            daily_kanji_readings.to_excel('C:/Users/ruben/Desktop/毎日の漢字.xlsx',"lecturas",index = False)
            book = load_workbook('C:/Users/ruben/Desktop/毎日の漢字.xlsx')
            writer = pd.ExcelWriter('C:/Users/ruben/Desktop/毎日の漢字.xlsx', engine='openpyxl') 
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            daily_kanji.to_excel(writer, "漢字, いみ",index = False)
            writer.save()
            os.system(r"start EXCEL.EXE C:\Users\ruben\Desktop\毎日の漢字.xlsx")
                    

              

ventana = Tk()
ventana.title('毎日の漢字プロガム')
icon = PhotoImage(file="A:/Rubén/Librerías/Documentos/Carpeta Compartida Linux/Python/毎日の漢字/日本語.png")
ventana.resizable(False,False)
ventana.iconphoto(False, icon, icon) 
ventana.config(background='white')


frame=Frame()
frame.pack()
frame.config(bg='white')
frame.config(width="610", height ="430")


image_1 = PhotoImage(file="A:/Rubén/Librerías/Documentos/Carpeta Compartida Linux/Python/毎日の漢字/b.png")
Label(frame, image= image_1,highlightthickness = 0,borderwidth = 0).place(x=0,y=0)
Label(frame,text = '毎日の漢字 ',background='white',fg='black',font=('Arial',18)).place(x=200,y=20)
Label(frame,text = '毎日の漢字数',background='white',fg='black',font=('Arial',12)).place(x=200,y=110)



num_kanjis = IntVar()
num_kanjis = Entry(frame,width=4, font=('Arial 10'))
num_kanjis.insert(0, "50")
num_kanjis.place(x=320,y=110)


new_kj = IntVar(value = 1)
excel_display = IntVar()
Checkbutton(frame, text='新漢字', background='white', highlightthickness=0,borderwidth =0,font=('Arial',11),variable = new_kj, onvalue = 1,offvalue = 0).place(x=200,y=160)
Checkbutton(frame, text='エクセルを開く', background='white', highlightthickness=0,borderwidth =0,font=('Arial',11),variable = excel_display, onvalue = 1,offvalue = 0).place(x=200,y=190)


Button(frame,text='引き金',command=kanjis).place(x=230,y=250)



ventana.mainloop()